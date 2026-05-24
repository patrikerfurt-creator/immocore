import io
import re
import uuid
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.cache import cache
from django.db import transaction

from apps.objekte.models import (
    Einheit, EinheitVerbrauch, Verteilerschluessel,
    VerteilerschluesselWert, Wirtschaftsjahr,
)
from apps.buchhaltung.models import VerteilerImportProtokoll

from .konstanten import (
    STAMM_VS_DIREKT, STAMM_VS_KOPF, VERBRAUCHS_VS_CODES, ALL_KNOWN_VS,
)

PREVIEW_TTL = 1800  # 30 Minuten


@dataclass
class ImportZeile:
    einheit_nr: str
    bezeichnung: str
    alter_wert: Decimal | None
    neuer_wert: Decimal | None
    status: str   # neu | geaendert | unveraendert | leer | ungueltig
    fehler: str | None = None
    warnung: str | None = None


@dataclass
class ParsedDatei:
    vs_code: str
    wj_jahr: int | None
    wj_id: str | None       # UUID-String des Wirtschaftsjahr-Objekts
    objekt_nr: str
    dateiname: str
    zeilen: list[ImportZeile] = field(default_factory=list)
    warnungen: list[str] = field(default_factory=list)
    fehler: list[str] = field(default_factory=list)

    @property
    def hat_fehler(self) -> bool:
        return bool(self.fehler) or any(z.fehler for z in self.zeilen)


class ParseError(ValueError):
    pass


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

_DATEINAME_STAMM_RE    = re.compile(r'^VS_(\w+)_(\d{3})\.xlsx$', re.IGNORECASE)
_DATEINAME_VERBRAUCH_RE = re.compile(r'^VS_(\w+)_(\d{3})_WJ_(\d{4})\.xlsx$', re.IGNORECASE)


def parse_vs_datei(file_bytes: bytes, dateiname: str, objekt) -> ParsedDatei:
    # 1. Dateiname parsen
    vs_code, wj_jahr = _parse_dateiname(dateiname)

    if vs_code in STAMM_VS_KOPF:
        raise ParseError(
            f"VS {vs_code} wird aus Einheit-Typ abgeleitet, Import nicht zulässig."
        )
    if vs_code not in ALL_KNOWN_VS:
        raise ParseError(f"VS-Code {vs_code} nicht bekannt.")

    # 2. WJ auflösen (nur Verbrauchs-VS)
    wj = None
    if vs_code in VERBRAUCHS_VS_CODES:
        if wj_jahr is None:
            raise ParseError(f"Wirtschaftsjahr für VS {vs_code} fehlt.")
        try:
            wj = Wirtschaftsjahr.objects.get(objekt=objekt, jahr=wj_jahr)
        except Wirtschaftsjahr.DoesNotExist:
            raise ParseError(f"Wirtschaftsjahr {wj_jahr} existiert nicht am Objekt.")
        if wj.status == 'abgeschlossen':
            raise ParseError(f"Wirtschaftsjahr {wj_jahr} ist abgeschlossen.")

    # 3. Workbook laden
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise ParseError("Datei ist kein gültiges .xlsx-Format.")

    if vs_code not in wb.sheetnames:
        raise ParseError(f"Tabellenblatt '{vs_code}' nicht gefunden.")

    ws = wb[vs_code]

    # 4. Header validieren
    header_objekt_nr = _header_wert(ws, 2)
    header_vs_code   = _header_wert(ws, 3)
    header_wj        = _header_wert(ws, 4)

    # B2 muss Objekt-Nr. enthalten
    if header_objekt_nr and not header_objekt_nr.startswith(objekt.objektnummer):
        raise ParseError(
            f"Objekt-Nr. in B2 ({header_objekt_nr!r}) stimmt nicht mit Objekt überein."
        )

    # B3 muss mit vs_code beginnen
    if header_vs_code and not header_vs_code.startswith(vs_code):
        raise ParseError(
            f"VS-Code in B3 ({header_vs_code!r}) stimmt nicht mit Dateiname ({vs_code}) überein."
        )

    # B4: bei Verbrauchs-VS muss WJ-Jahr übereinstimmen
    if vs_code in VERBRAUCHS_VS_CODES and header_wj:
        header_wj_clean = str(header_wj).strip()
        if header_wj_clean != str(wj_jahr):
            raise ParseError(
                f"Wirtschaftsjahr in B4 ({header_wj_clean}) stimmt nicht mit Dateiname ({wj_jahr}) überein."
            )

    wb.close()

    # 5. Datenzeilen einlesen (ab Zeile 9, max. 500)
    wb2 = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws2 = wb2[vs_code]

    roh_zeilen: list[tuple[str, str, str | float | None]] = []
    for row in ws2.iter_rows(min_row=9, values_only=True):
        if len(roh_zeilen) >= 500:
            raise ParseError("Maximal 500 Einheiten je Datei.")
        einheit_nr_raw = row[0] if len(row) > 0 else None
        bezeichnung    = str(row[1]) if len(row) > 1 and row[1] is not None else ''
        wert_raw       = row[2] if len(row) > 2 else None
        if einheit_nr_raw is None:
            continue
        roh_zeilen.append((str(einheit_nr_raw).strip(), bezeichnung, wert_raw))

    wb2.close()

    # 6. Duplikat-Check
    nrs = [z[0] for z in roh_zeilen]
    if len(nrs) != len(set(nrs)):
        duplikate = [nr for nr in set(nrs) if nrs.count(nr) > 1]
        raise ParseError(f"Einheit-Nr. doppelt in der Datei: {', '.join(duplikate)}")

    # 7. Alle Einheiten des Objekts aufbauen
    einheit_map = {
        e.einheit_nr: e
        for e in Einheit.objects.filter(objekt=objekt)
    }

    # 8. Aktuelle Werte laden
    akt_werte = _lade_aktuelle_werte(objekt, wj, vs_code)

    # 9. Zeilen mit Diff aufbauen
    parsed = ParsedDatei(
        vs_code=vs_code,
        wj_jahr=wj_jahr,
        wj_id=str(wj.id) if wj else None,
        objekt_nr=objekt.objektnummer,
        dateiname=dateiname,
    )

    for einheit_nr, bezeichnung, wert_raw in roh_zeilen:
        zeile = ImportZeile(
            einheit_nr=einheit_nr,
            bezeichnung=bezeichnung,
            alter_wert=akt_werte.get(einheit_nr),
            neuer_wert=None,
            status='unveraendert',
        )

        if einheit_nr not in einheit_map:
            zeile.fehler = f"Einheit {einheit_nr} existiert nicht am Objekt."
            zeile.status = 'ungueltig'
            parsed.zeilen.append(zeile)
            continue

        neuer_wert = _parse_wert(wert_raw)
        zeile.neuer_wert = neuer_wert

        if neuer_wert is None:
            zeile.status = 'leer'
        elif zeile.alter_wert is None:
            zeile.status = 'neu'
        elif neuer_wert != zeile.alter_wert:
            zeile.status = 'geaendert'
        else:
            zeile.status = 'unveraendert'

        if neuer_wert is not None and neuer_wert < 0:
            zeile.warnung = "Wert ist negativ."

        parsed.zeilen.append(zeile)

    # Einheiten die im Objekt existieren, aber nicht in der Datei
    datei_nrs = {z.einheit_nr for z in parsed.zeilen}
    fehlende = [nr for nr in einheit_map if nr not in datei_nrs]
    if fehlende:
        parsed.warnungen.append(
            f"{len(fehlende)} Einheit(en) nicht in Datei — bleiben unverändert: {', '.join(sorted(fehlende))}"
        )

    if vs_code in STAMM_VS_DIREKT:
        parsed.warnungen.insert(0,
            "Achtung: Diese Werte überschreiben Einheit-Stammdaten (Verteilerschlüssel-Werte)."
        )

    return parsed


def erstelle_preview(parsed: ParsedDatei, objekt, user) -> tuple[str, dict]:
    token = str(uuid.uuid4())
    daten = {
        'objekt_id': str(objekt.id),
        'vs_code': parsed.vs_code,
        'wj_id': parsed.wj_id,
        'dateiname': parsed.dateiname,
        'zeilen': [
            {
                'einheit_nr':  z.einheit_nr,
                'bezeichnung': z.bezeichnung,
                'alter_wert':  str(z.alter_wert) if z.alter_wert is not None else None,
                'neuer_wert':  str(z.neuer_wert) if z.neuer_wert is not None else None,
                'status':      z.status,
                'fehler':      z.fehler,
                'warnung':     z.warnung,
            }
            for z in parsed.zeilen
        ],
        'warnungen': parsed.warnungen,
        'fehler': parsed.fehler,
    }
    cache.set(f'verteiler_preview:{token}', daten, timeout=PREVIEW_TTL)

    zusammenfassung = _zusammenfassung(parsed.zeilen)
    vorschau = {
        'preview_token': token,
        'vs_code': parsed.vs_code,
        'wj_jahr': parsed.wj_jahr,
        'dateiname': parsed.dateiname,
        'zeilen': daten['zeilen'],
        'warnungen': parsed.warnungen,
        'fehler': parsed.fehler,
        'hat_fehler': parsed.hat_fehler,
        'zusammenfassung': zusammenfassung,
    }
    return token, vorschau


# ---------------------------------------------------------------------------
# Commit
# ---------------------------------------------------------------------------

@transaction.atomic
def commit_verteiler_import(preview_token: str, user) -> dict:
    daten = cache.get(f'verteiler_preview:{preview_token}')
    if not daten:
        raise ValueError("Vorschau abgelaufen, bitte Datei erneut hochladen.")

    from apps.objekte.models import Objekt
    objekt = Objekt.objects.select_for_update().get(id=daten['objekt_id'])
    vs_code = daten['vs_code']
    wj_id   = daten.get('wj_id')

    zeilen = daten['zeilen']

    # Fehlercheck: kein Commit bei ungültigen Zeilen
    if any(z['fehler'] for z in zeilen):
        raise ValueError("Datei enthält ungültige Zeilen — bitte korrigieren und erneut hochladen.")

    geaendert = 0

    if vs_code in STAMM_VS_DIREKT:
        vs_obj = Verteilerschluessel.objects.get(objekt=objekt, schluessel=vs_code)
        for zeile in zeilen:
            if zeile['status'] in ('neu', 'geaendert') and zeile['neuer_wert'] is not None:
                einheit = Einheit.objects.get(objekt=objekt, einheit_nr=zeile['einheit_nr'])
                VerteilerschluesselWert.objects.update_or_create(
                    schluessel=vs_obj,
                    einheit=einheit,
                    wirtschaftsjahr=0,
                    defaults={'wert': Decimal(zeile['neuer_wert']), 'quelle': 'manuell'},
                )
                geaendert += 1

    elif vs_code in VERBRAUCHS_VS_CODES:
        wj = Wirtschaftsjahr.objects.get(id=wj_id)
        if wj.status == 'abgeschlossen':
            raise ValueError(f"Wirtschaftsjahr {wj.jahr} ist abgeschlossen.")

        for zeile in zeilen:
            if zeile['status'] in ('neu', 'geaendert', 'leer'):
                einheit = Einheit.objects.get(objekt=objekt, einheit_nr=zeile['einheit_nr'])
                neuer_wert = Decimal(zeile['neuer_wert']) if zeile['neuer_wert'] is not None else None
                EinheitVerbrauch.objects.update_or_create(
                    wirtschaftsjahr=wj,
                    einheit=einheit,
                    vs_code=vs_code,
                    defaults={'wert': neuer_wert, 'quelle': 'manuell'},
                )
                geaendert += 1

    else:
        raise ValueError(f"VS-Code {vs_code} nicht importierbar.")

    VerteilerImportProtokoll.objects.create(
        objekt=objekt,
        wirtschaftsjahr_id=wj_id,
        vs_code=vs_code,
        dateiname=daten['dateiname'],
        anzahl_aktualisiert=geaendert,
        importiert_von=user,
    )

    cache.delete(f'verteiler_preview:{preview_token}')
    return {'anzahl_aktualisiert': geaendert}


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------

def _parse_dateiname(dateiname: str) -> tuple[str, int | None]:
    name = dateiname.strip()
    m = _DATEINAME_VERBRAUCH_RE.match(name)
    if m:
        return m.group(2), int(m.group(3))
    m = _DATEINAME_STAMM_RE.match(name)
    if m:
        return m.group(2), None
    raise ParseError(
        f"Dateiname '{name}' entspricht nicht dem Schema "
        f"VS_{{OBJEKT_NR}}_{{VS_CODE}}[_WJ_{{JAHR}}].xlsx"
    )


def _header_wert(ws, row: int) -> str | None:
    cell = ws.cell(row=row, column=2)
    val = cell.value
    return str(val).strip() if val is not None else None


def _parse_wert(raw) -> Decimal | None:
    if raw is None or str(raw).strip() == '':
        return None
    try:
        normalized = str(raw).replace(',', '.').strip()
        return Decimal(normalized)
    except (InvalidOperation, ValueError):
        return None


def _lade_aktuelle_werte(objekt, wj, vs_code: str) -> dict:
    """Gibt {einheit_nr: Decimal|None} zurück."""
    result = {}
    if vs_code in STAMM_VS_DIREKT:
        vs_obj = Verteilerschluessel.objects.filter(objekt=objekt, schluessel=vs_code).first()
        if vs_obj:
            for vsw in VerteilerschluesselWert.objects.filter(
                schluessel=vs_obj, wirtschaftsjahr=0
            ).select_related('einheit'):
                result[vsw.einheit.einheit_nr] = vsw.wert
    elif vs_code in VERBRAUCHS_VS_CODES and wj:
        for ev in EinheitVerbrauch.objects.filter(
            wirtschaftsjahr=wj, einheit__objekt=objekt, vs_code=vs_code
        ).select_related('einheit'):
            result[ev.einheit.einheit_nr] = ev.wert
    return result


def _zusammenfassung(zeilen: list[ImportZeile | dict]) -> dict:
    counts = {'neu': 0, 'geaendert': 0, 'unveraendert': 0, 'leer': 0, 'ungueltig': 0}
    for z in zeilen:
        s = z['status'] if isinstance(z, dict) else z.status
        counts[s] = counts.get(s, 0) + 1
    return counts
