import io
import zipfile
from dataclasses import dataclass, field
from decimal import Decimal

import openpyxl
from django.utils import timezone
from openpyxl.styles import Font, PatternFill, Alignment

from apps.objekte.models import (
    Einheit, EinheitVerbrauch, Verteilerschluessel,
    VerteilerschluesselWert, Wirtschaftsjahr,
)
from apps.konten.models import Konto, KontoVerteilerSchluessel
from django.db.models import Case, When, Value, IntegerField

from .konstanten import (
    STAMM_VS_DIREKT, STAMM_VS_KOPF, VERBRAUCHS_VS_CODES,
    ALL_KNOWN_VS, STAMM_VS_MAASSEINHEIT, ZELL_FORMAT,
)


@dataclass
class VsInfo:
    code: str
    bezeichnung: str
    kategorie: str  # stamm_direkt | stamm_kopf | verbrauch
    wirtschaftsjahre: list = field(default_factory=list)  # nur bei verbrauch


@dataclass
class VsExportRequest:
    code: str
    wj_id: str | None = None  # UUID-String; None für Stamm-VS


def ermittle_aktive_vs(objekt) -> list[VsInfo]:
    """Vereinigungsmenge der vs_codes aus beiden VS-Quellen für alle WJ des Objekts.

    Quelle 1: KontoVerteilerSchluessel (Join-Tabelle, neuere Kontenrahmen-Importe)
    Quelle 2: Konto.verteilerschluessel (CharField, ältere / importierte Kontenrahmen)
    """
    from_kvs = set(
        KontoVerteilerSchluessel.objects
        .filter(konto__wirtschaftsjahr__objekt=objekt)
        .values_list('vs_code', flat=True)
        .distinct()
    )
    from_konto = set(
        Konto.objects
        .filter(wirtschaftsjahr__objekt=objekt)
        .exclude(verteilerschluessel__isnull=True)
        .exclude(verteilerschluessel='')
        .values_list('verteilerschluessel', flat=True)
        .distinct()
    )
    aktive_codes = sorted(from_kvs | from_konto)

    wj_liste = list(
        Wirtschaftsjahr.objects.filter(objekt=objekt).order_by('jahr')
    )

    result = []
    for code in aktive_codes:
        if code not in ALL_KNOWN_VS:
            continue
        vs_obj = Verteilerschluessel.objects.filter(objekt=objekt, schluessel=code).first()
        bezeichnung = vs_obj.bezeichnung if vs_obj else f"VS {code}"

        kategorie = (
            'stamm_direkt' if code in STAMM_VS_DIREKT else
            'stamm_kopf'   if code in STAMM_VS_KOPF   else
            'verbrauch'
        )
        entry = VsInfo(code=code, bezeichnung=bezeichnung, kategorie=kategorie)
        if code in VERBRAUCHS_VS_CODES:
            entry.wirtschaftsjahre = [
                {'id': str(wj.id), 'jahr': wj.jahr, 'status': wj.status}
                for wj in wj_liste
            ]
        result.append(entry)

    return result


def export_verteiler_zip(objekt, anforderungen: list[VsExportRequest]) -> bytes:
    aeltestes_wj = Wirtschaftsjahr.objects.filter(objekt=objekt).order_by('jahr').first()

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for req in anforderungen:
            code = req.code
            if code in STAMM_VS_DIREKT | STAMM_VS_KOPF:
                wj = aeltestes_wj
                dateiname = f"VS_{objekt.objektnummer}_{code}.xlsx"
            elif code in VERBRAUCHS_VS_CODES:
                wj = Wirtschaftsjahr.objects.get(id=req.wj_id, objekt=objekt)
                dateiname = f"VS_{objekt.objektnummer}_{code}_WJ_{wj.jahr}.xlsx"
            else:
                raise ValueError(f"Unbekannter VS-Code: {code}")
            zf.writestr(dateiname, baue_vs_excel(objekt, wj, code))

    return zip_buffer.getvalue()


def baue_vs_excel(objekt, wj, vs_code: str) -> bytes:
    einheiten = _einheiten_sortiert(objekt)
    vs_obj = Verteilerschluessel.objects.filter(objekt=objekt, schluessel=vs_code).first()
    vs_bezeichnung = vs_obj.bezeichnung if vs_obj else f"VS {vs_code}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = vs_code

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    bold = Font(bold=True)

    def _meta(row, label, value):
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        for c in range(1, 3):
            ws.cell(row=row, column=c).fill = header_fill

    wj_text = str(wj.jahr) if vs_code in VERBRAUCHS_VS_CODES else "— (objektweit)"

    _meta(1, "IMMOCORE Verteilerschlüssel-Import", None)
    ws.merge_cells('A1:C1')
    ws['A1'].font = bold
    ws['A1'].fill = header_fill

    _meta(2, "Objekt:", f"{objekt.objektnummer} — {objekt.bezeichnung}")
    _meta(3, "VS-Code:", f"{vs_code} — {vs_bezeichnung}")
    _meta(4, "Wirtschaftsjahr:", wj_text)
    _meta(5, "Maßeinheit:", _ermittle_einheit_text(objekt, wj, vs_code))
    _meta(6, "Erzeugt am:", timezone.localtime(timezone.now()).strftime("%d.%m.%Y %H:%M"))

    # Zeile 7 leer, Zeile 8 Spaltenköpfe
    for col, header in enumerate(["einheit_nr", "bezeichnung", "wert"], start=1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = bold

    # Datenzeilen ab Zeile 9
    zell_fmt = ZELL_FORMAT.get(vs_code, '#,##0.00')
    wert_map = _lade_wert_map(objekt, wj, vs_code, einheiten)

    for i, einheit in enumerate(einheiten, start=9):
        ws.cell(row=i, column=1, value=einheit.einheit_nr)
        ws.cell(row=i, column=2, value=einheit.lage)
        wert = wert_map.get(einheit.id)
        wert_cell = ws.cell(row=i, column=3, value=float(wert) if wert is not None else None)
        wert_cell.number_format = zell_fmt

    # Spaltenbreiten
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 16

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------

def _einheiten_sortiert(objekt):
    return list(
        Einheit.objects.filter(objekt=objekt).annotate(
            typ_order=Case(
                When(einheit_typ='Wohnung',    then=Value(0)),
                When(einheit_typ='Gewerbe',    then=Value(1)),
                When(einheit_typ='Stellplatz', then=Value(2)),
                When(einheit_typ='Sonstiges',  then=Value(3)),
                default=Value(4),
                output_field=IntegerField(),
            )
        ).order_by('typ_order', 'einheit_nr')
    )


def _ermittle_einheit_text(objekt, wj, vs_code: str) -> str:
    if vs_code in STAMM_VS_MAASSEINHEIT:
        return STAMM_VS_MAASSEINHEIT[vs_code]
    ev = (
        EinheitVerbrauch.objects
        .filter(wirtschaftsjahr=wj, einheit__objekt=objekt, vs_code=vs_code)
        .exclude(einheit_text='')
        .first()
    )
    return ev.einheit_text if ev else ''


def _lade_wert_map(objekt, wj, vs_code: str, einheiten) -> dict:
    """Gibt {einheit.id: Decimal|None} zurück."""
    result = {}

    if vs_code in STAMM_VS_DIREKT:
        vs_obj = Verteilerschluessel.objects.filter(objekt=objekt, schluessel=vs_code).first()
        if vs_obj:
            for vsw in VerteilerschluesselWert.objects.filter(
                schluessel=vs_obj, wirtschaftsjahr=0
            ).select_related('einheit'):
                result[vsw.einheit_id] = vsw.wert
        return result

    if vs_code == '030':
        return {e.id: Decimal('1') for e in einheiten}

    if vs_code == '031':
        return {e.id: Decimal('1') if e.einheit_typ == 'Wohnung' else Decimal('0') for e in einheiten}

    if vs_code == '032':
        return {e.id: Decimal('1') if e.einheit_typ == 'Stellplatz' else Decimal('0') for e in einheiten}

    if vs_code in VERBRAUCHS_VS_CODES:
        for ev in EinheitVerbrauch.objects.filter(
            wirtschaftsjahr=wj, einheit__objekt=objekt, vs_code=vs_code
        ).select_related('einheit'):
            result[ev.einheit_id] = ev.wert
        return result

    return result
