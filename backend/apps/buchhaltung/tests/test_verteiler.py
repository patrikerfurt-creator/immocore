"""
Tests: Verteilerschlüssel Export & Import (Spec v1.0)
"""
import io
import zipfile
from datetime import date
from decimal import Decimal

import openpyxl
from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from rest_framework.test import APIClient

from apps.buchhaltung.models import VerteilerImportProtokoll
from apps.buchhaltung.services.verteiler.export_service import (
    baue_vs_excel, export_verteiler_zip, ermittle_aktive_vs, VsExportRequest,
)
from apps.buchhaltung.services.verteiler.import_service import (
    parse_vs_datei, erstelle_preview, commit_verteiler_import, ParseError,
)
from apps.objekte.models import (
    Einheit, EinheitVerbrauch, Objekt, Verteilerschluessel,
    VerteilerschluesselWert, Wirtschaftsjahr,
)
from apps.konten.models import Konto, KontoVerteilerSchluessel

User = get_user_model()


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _user(username='vs-tester'):
    u, _ = User.objects.get_or_create(username=username, defaults={'is_staff': True})
    return u


def _objekt(nr='100001'):
    return Objekt.objects.create(
        objektnummer=nr,
        objekt_typ='WEG',
        bezeichnung='VS-Test-Objekt',
        strasse='Teststr. 1',
        plz='60311',
        ort='Frankfurt',
        verwaltung_seit=date(2020, 1, 1),
        glaeubiger_id='DE98ZZZ09999999999',
    )


def _wj(objekt, jahr=2025):
    return Wirtschaftsjahr.objects.create(
        objekt=objekt, jahr=jahr, beginn_monat=1, status='offen',
    )


def _einheit(objekt, nr, typ='Wohnung', lage='EG links'):
    return Einheit.objects.create(
        objekt=objekt, einheit_nr=nr, einheit_typ=typ, lage=lage,
    )


def _vs(objekt, code, bezeichnung='Test-VS'):
    return Verteilerschluessel.objects.create(
        objekt=objekt, schluessel=code, bezeichnung=bezeichnung,
    )


def _vsw(vs_obj, einheit, wert, wj=0):
    return VerteilerschluesselWert.objects.create(
        schluessel=vs_obj, einheit=einheit, wirtschaftsjahr=wj, wert=wert,
    )


def _konto(wj, nr='4100', vs_code='001'):
    konto = Konto.objects.create(wirtschaftsjahr=wj, kontonummer=nr, kontoname=f'Konto {nr}')
    KontoVerteilerSchluessel.objects.create(konto=konto, vs_code=vs_code, gueltig_ab=date(2025, 1, 1))
    return konto


# ---------------------------------------------------------------------------
# Export-Tests
# ---------------------------------------------------------------------------

class ExportStammVSTest(TestCase):
    def setUp(self):
        self.obj = _objekt()
        self.wj  = _wj(self.obj)
        self.e1  = _einheit(self.obj, 'WE01', 'Wohnung', 'EG links')
        self.e2  = _einheit(self.obj, 'WE02', 'Wohnung', '1.OG rechts')
        self.vs001 = _vs(self.obj, '001', 'Wohnfläche')
        _vsw(self.vs001, self.e1, Decimal('65.50'))
        _vsw(self.vs001, self.e2, Decimal('72.00'))

    def test_export_001_befuellung(self):
        xlsx = baue_vs_excel(self.obj, self.wj, '001')
        wb = openpyxl.load_workbook(io.BytesIO(xlsx), data_only=True)
        ws = wb['001']
        self.assertEqual(ws['B3'].value, '001 — Wohnfläche')
        self.assertEqual(ws['B4'].value, '— (objektweit)')
        self.assertEqual(ws['B5'].value, 'm²')
        # Datenzeilen ab 9 — Reihenfolge nach einheit_typ, einheit_nr
        werte = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=3).value for r in range(9, 11)}
        self.assertAlmostEqual(float(werte['WE01']), 65.50)
        self.assertAlmostEqual(float(werte['WE02']), 72.00)

    def test_export_031_kopf(self):
        _vs(self.obj, '031', 'Anzahl Wohnungen')
        e3 = _einheit(self.obj, 'ST01', 'Stellplatz', 'Tiefgarage')
        xlsx = baue_vs_excel(self.obj, self.wj, '031')
        wb = openpyxl.load_workbook(io.BytesIO(xlsx), data_only=True)
        ws = wb['031']
        werte = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=3).value
                 for r in range(9, 13) if ws.cell(row=r, column=1).value}
        self.assertEqual(werte.get('WE01'), 1)
        self.assertEqual(werte.get('WE02'), 1)
        self.assertEqual(werte.get('ST01'), 0)


class ExportVerbrauchVSTest(TestCase):
    def setUp(self):
        self.obj = _objekt('100002')
        self.wj  = _wj(self.obj, 2025)
        self.e1  = _einheit(self.obj, 'WE01')
        self.e2  = _einheit(self.obj, 'WE02')
        self.vs140 = _vs(self.obj, '140', 'Heizung')

    def test_export_140_mit_werten(self):
        EinheitVerbrauch.objects.create(
            wirtschaftsjahr=self.wj, einheit=self.e1,
            vs_code='140', wert=Decimal('1234.5678'), einheit_text='kWh',
        )
        xlsx = baue_vs_excel(self.obj, self.wj, '140')
        wb = openpyxl.load_workbook(io.BytesIO(xlsx), data_only=True)
        ws = wb['140']
        self.assertEqual(ws['B4'].value, '2025')
        self.assertEqual(ws['B5'].value, 'kWh')
        werte = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=3).value
                 for r in range(9, 12) if ws.cell(row=r, column=1).value}
        self.assertAlmostEqual(float(werte['WE01']), 1234.5678, places=4)
        self.assertIsNone(werte.get('WE02'))

    def test_export_140_null_bleibt_leer(self):
        xlsx = baue_vs_excel(self.obj, self.wj, '140')
        wb = openpyxl.load_workbook(io.BytesIO(xlsx), data_only=True)
        ws = wb['140']
        werte = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=3).value
                 for r in range(9, 12) if ws.cell(row=r, column=1).value}
        self.assertIsNone(werte.get('WE01'))
        self.assertIsNone(werte.get('WE02'))


class ExportZipTest(TestCase):
    def setUp(self):
        self.obj = _objekt('100003')
        self.wj  = _wj(self.obj, 2025)
        _einheit(self.obj, 'WE01')
        _vs(self.obj, '001', 'Wohnfläche')
        _vs(self.obj, '010', 'MEA')
        _vs(self.obj, '140', 'Heizung')

    def test_zip_enthaelt_3_dateien(self):
        anforderungen = [
            VsExportRequest(code='001'),
            VsExportRequest(code='010'),
            VsExportRequest(code='140', wj_id=str(self.wj.id)),
        ]
        zip_bytes = export_verteiler_zip(self.obj, anforderungen)
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            namen = zf.namelist()
        self.assertEqual(len(namen), 3)
        self.assertIn(f'VS_100003_001.xlsx', namen)
        self.assertIn(f'VS_100003_010.xlsx', namen)
        self.assertIn(f'VS_100003_140_WJ_2025.xlsx', namen)


# ---------------------------------------------------------------------------
# Import-Tests
# ---------------------------------------------------------------------------

@override_settings(CACHES={'default': {'BACKEND': 'django.core.cache.backends.locmem.LocMemCache'}})
class ImportStammVSTest(TestCase):
    def setUp(self):
        self.user = _user()
        self.obj  = _objekt('100004')
        self.wj   = _wj(self.obj, 2025)
        self.e1   = _einheit(self.obj, 'WE01', lage='EG links')
        self.e2   = _einheit(self.obj, 'WE02', lage='1.OG')
        self.vs001 = _vs(self.obj, '001', 'Wohnfläche')

    def _baue_xlsx(self, zeilen: list[tuple], vs_code='001', objekt_nr='100004', wj_jahr=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = vs_code
        ws['A1'] = 'IMMOCORE Verteilerschlüssel-Import'
        ws['A2'] = 'Objekt:';    ws['B2'] = f'{objekt_nr} — Test'
        ws['A3'] = 'VS-Code:';   ws['B3'] = f'{vs_code} — Test-VS'
        ws['A4'] = 'Wirtschaftsjahr:'; ws['B4'] = str(wj_jahr) if wj_jahr else '— (objektweit)'
        ws['A5'] = 'Maßeinheit:'; ws['B5'] = 'm²'
        ws['A6'] = 'Erzeugt am:'; ws['B6'] = '24.05.2026 14:30'
        ws['A8'] = 'einheit_nr'; ws['B8'] = 'bezeichnung'; ws['C8'] = 'wert'
        for i, (nr, bez, wert) in enumerate(zeilen, start=9):
            ws.cell(row=i, column=1, value=nr)
            ws.cell(row=i, column=2, value=bez)
            ws.cell(row=i, column=3, value=wert)
        buf = io.BytesIO(); wb.save(buf)
        return buf.getvalue()

    def test_import_001_schreibt_vsw(self):
        xlsx = self._baue_xlsx([('WE01', 'EG links', 65.5), ('WE02', '1.OG', 72.0)])
        parsed = parse_vs_datei(xlsx, 'VS_100004_001.xlsx', self.obj)
        self.assertFalse(parsed.hat_fehler)
        self.assertEqual(len([z for z in parsed.zeilen if z.status == 'neu']), 2)

        token, _ = erstelle_preview(parsed, self.obj, self.user)
        result = commit_verteiler_import(token, self.user)
        self.assertEqual(result['anzahl_aktualisiert'], 2)

        vsw = VerteilerschluesselWert.objects.get(schluessel=self.vs001, einheit=self.e1, wirtschaftsjahr=0)
        self.assertEqual(vsw.wert, Decimal('65.5'))

    def test_import_leere_zelle_unveraendert_lassen(self):
        _vsw(self.vs001, self.e1, Decimal('65.5'))
        xlsx = self._baue_xlsx([('WE01', 'EG links', None), ('WE02', '1.OG', 72.0)])
        parsed = parse_vs_datei(xlsx, 'VS_100004_001.xlsx', self.obj)
        e1_zeile = next(z for z in parsed.zeilen if z.einheit_nr == 'WE01')
        self.assertEqual(e1_zeile.status, 'leer')

        token, _ = erstelle_preview(parsed, self.obj, self.user)
        commit_verteiler_import(token, self.user)

        vsw = VerteilerschluesselWert.objects.get(schluessel=self.vs001, einheit=self.e1, wirtschaftsjahr=0)
        self.assertEqual(vsw.wert, Decimal('65.5'))  # unverändert

    def test_warn_banner_stamm_vs(self):
        xlsx = self._baue_xlsx([('WE01', 'EG links', 65.5)])
        parsed = parse_vs_datei(xlsx, 'VS_100004_001.xlsx', self.obj)
        self.assertTrue(any('überschreiben' in w for w in parsed.warnungen))

    def test_audit_eintrag(self):
        xlsx = self._baue_xlsx([('WE01', 'EG links', 65.5)])
        parsed = parse_vs_datei(xlsx, 'VS_100004_001.xlsx', self.obj)
        token, _ = erstelle_preview(parsed, self.obj, self.user)
        commit_verteiler_import(token, self.user)
        p = VerteilerImportProtokoll.objects.filter(objekt=self.obj, vs_code='001').first()
        self.assertIsNotNone(p)
        self.assertEqual(p.importiert_von, self.user)
        self.assertIsNone(p.wirtschaftsjahr)


@override_settings(CACHES={'default': {'BACKEND': 'django.core.cache.backends.locmem.LocMemCache'}})
class ImportVerbrauchVSTest(TestCase):
    def setUp(self):
        self.user = _user('vs-tester2')
        self.obj  = _objekt('100005')
        self.wj   = _wj(self.obj, 2025)
        self.e1   = _einheit(self.obj, 'WE01', lage='EG links')
        self.e2   = _einheit(self.obj, 'WE02', lage='1.OG')
        _vs(self.obj, '140', 'Heizung')

    def _baue_xlsx(self, zeilen):
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = '140'
        ws['A1'] = 'IMMOCORE Verteilerschlüssel-Import'
        ws['A2'] = 'Objekt:';    ws['B2'] = '100005 — Test'
        ws['A3'] = 'VS-Code:';   ws['B3'] = '140 — Heizung'
        ws['A4'] = 'Wirtschaftsjahr:'; ws['B4'] = '2025'
        ws['A5'] = 'Maßeinheit:'; ws['B5'] = 'kWh'
        ws['A6'] = 'Erzeugt am:'; ws['B6'] = '24.05.2026 14:30'
        ws['A8'] = 'einheit_nr'; ws['B8'] = 'bezeichnung'; ws['C8'] = 'wert'
        for i, (nr, bez, wert) in enumerate(zeilen, start=9):
            ws.cell(row=i, column=1, value=nr)
            ws.cell(row=i, column=2, value=bez)
            ws.cell(row=i, column=3, value=wert)
        buf = io.BytesIO(); wb.save(buf)
        return buf.getvalue()

    def test_import_140_schreibt_einheit_verbrauch(self):
        xlsx = self._baue_xlsx([('WE01', 'EG links', 1234.5678), ('WE02', '1.OG', 987.0)])
        parsed = parse_vs_datei(xlsx, 'VS_100005_140_WJ_2025.xlsx', self.obj)
        token, _ = erstelle_preview(parsed, self.obj, self.user)
        result = commit_verteiler_import(token, self.user)
        self.assertEqual(result['anzahl_aktualisiert'], 2)
        ev = EinheitVerbrauch.objects.get(wirtschaftsjahr=self.wj, einheit=self.e1, vs_code='140')
        self.assertEqual(ev.wert, Decimal('1234.5678'))

    def test_leere_zelle_setzt_wert_null(self):
        EinheitVerbrauch.objects.create(
            wirtschaftsjahr=self.wj, einheit=self.e1, vs_code='140', wert=Decimal('500')
        )
        xlsx = self._baue_xlsx([('WE01', 'EG links', None), ('WE02', '1.OG', 987.0)])
        parsed = parse_vs_datei(xlsx, 'VS_100005_140_WJ_2025.xlsx', self.obj)
        token, _ = erstelle_preview(parsed, self.obj, self.user)
        commit_verteiler_import(token, self.user)
        ev = EinheitVerbrauch.objects.get(wirtschaftsjahr=self.wj, einheit=self.e1, vs_code='140')
        self.assertIsNone(ev.wert)

    def test_audit_eintrag_mit_wj(self):
        xlsx = self._baue_xlsx([('WE01', 'EG links', 1000.0)])
        parsed = parse_vs_datei(xlsx, 'VS_100005_140_WJ_2025.xlsx', self.obj)
        token, _ = erstelle_preview(parsed, self.obj, self.user)
        commit_verteiler_import(token, self.user)
        p = VerteilerImportProtokoll.objects.filter(objekt=self.obj, vs_code='140').first()
        self.assertIsNotNone(p)
        self.assertEqual(p.wirtschaftsjahr, self.wj)


# ---------------------------------------------------------------------------
# Validierungs-Tests
# ---------------------------------------------------------------------------

class ParseValidierungTest(TestCase):
    def setUp(self):
        self.obj = _objekt('100006')
        self.wj_offen = _wj(self.obj, 2025)
        self.wj_abg   = Wirtschaftsjahr.objects.create(
            objekt=self.obj, jahr=2024, beginn_monat=1, status='abgeschlossen',
        )
        _einheit(self.obj, 'WE01')
        _vs(self.obj, '031', 'Anzahl Wohnungen')

    def _xlsx(self, vs_code, wj_jahr=None, zeilen=None):
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = vs_code
        ws['A1'] = 'IMMOCORE Verteilerschlüssel-Import'
        ws['A2'] = 'Objekt:';    ws['B2'] = f'100006 — Test'
        ws['A3'] = 'VS-Code:';   ws['B3'] = f'{vs_code} — Test'
        ws['A4'] = 'Wirtschaftsjahr:'; ws['B4'] = str(wj_jahr) if wj_jahr else '— (objektweit)'
        ws['A5'] = 'Maßeinheit:'; ws['B5'] = ''
        ws['A6'] = 'Erzeugt am:'; ws['B6'] = ''
        ws['A8'] = 'einheit_nr'; ws['B8'] = 'bezeichnung'; ws['C8'] = 'wert'
        for i, row in enumerate(zeilen or [('WE01', '', 1)], start=9):
            for col, val in enumerate(row, start=1):
                ws.cell(row=i, column=col, value=val)
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    def test_031_wird_abgelehnt(self):
        xlsx = self._xlsx('031')
        with self.assertRaises(ParseError) as ctx:
            parse_vs_datei(xlsx, 'VS_100006_031.xlsx', self.obj)
        self.assertIn('abgeleitet', str(ctx.exception))

    def test_abgeschlossenes_wj_abgelehnt(self):
        xlsx = self._xlsx('140', wj_jahr=2024)
        with self.assertRaises(ParseError) as ctx:
            parse_vs_datei(xlsx, 'VS_100006_140_WJ_2024.xlsx', self.obj)
        self.assertIn('abgeschlossen', str(ctx.exception))

    def test_dateiname_header_mismatch(self):
        xlsx = self._xlsx('001')
        # Dateiname sagt 001, aber Header B3 sagt 010
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        wb['001']['B3'] = '010 — MEA'
        buf = io.BytesIO(); wb.save(buf)
        with self.assertRaises(ParseError) as ctx:
            parse_vs_datei(buf.getvalue(), 'VS_100006_001.xlsx', self.obj)
        self.assertIn('stimmt nicht', str(ctx.exception))

    def test_unbekannte_einheit_nr(self):
        xlsx = self._xlsx('001', zeilen=[('WE01', '', 65.5), ('WE99', '', 50.0)])
        parsed = parse_vs_datei(xlsx, 'VS_100006_001.xlsx', self.obj)
        ungueltig = [z for z in parsed.zeilen if z.status == 'ungueltig']
        self.assertEqual(len(ungueltig), 1)
        self.assertEqual(ungueltig[0].einheit_nr, 'WE99')

    def test_falsches_dateiname_schema(self):
        with self.assertRaises(ParseError):
            parse_vs_datei(b'x', 'FALSCH.xlsx', self.obj)


# ---------------------------------------------------------------------------
# API-Tests
# ---------------------------------------------------------------------------

@override_settings(CACHES={'default': {'BACKEND': 'django.core.cache.backends.locmem.LocMemCache'}})
class VerteilerAPITest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = _user('api-tester')
        self.client.force_authenticate(user=self.user)
        self.obj = _objekt('100007')
        self.wj  = _wj(self.obj, 2025)
        self.e1  = _einheit(self.obj, 'WE01')
        _vs(self.obj, '001', 'Wohnfläche')
        _konto(self.wj, '4200', '001')

    def test_aktive_vs_liste(self):
        resp = self.client.get(f'/api/v1/objekte/{self.obj.id}/verteiler/aktive-vs/')
        self.assertEqual(resp.status_code, 200)
        codes = [v['code'] for v in resp.json()]
        self.assertIn('001', codes)

    def test_export_liefert_zip(self):
        resp = self.client.post(
            f'/api/v1/objekte/{self.obj.id}/verteiler/export/',
            {'vs_codes': [{'code': '001'}]},
            format='json',
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/zip')

    def test_export_ohne_auswahl_400(self):
        resp = self.client.post(
            f'/api/v1/objekte/{self.obj.id}/verteiler/export/',
            {'vs_codes': []},
            format='json',
        )
        self.assertEqual(resp.status_code, 400)

    def test_import_preview_und_commit(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = '001'
        ws['A1'] = 'IMMOCORE Verteilerschlüssel-Import'
        ws['A2'] = 'Objekt:'; ws['B2'] = '100007 — Test'
        ws['A3'] = 'VS-Code:'; ws['B3'] = '001 — Wohnfläche'
        ws['A4'] = 'Wirtschaftsjahr:'; ws['B4'] = '— (objektweit)'
        ws['A5'] = 'Maßeinheit:'; ws['B5'] = 'm²'
        ws['A6'] = 'Erzeugt am:'; ws['B6'] = ''
        ws['A8'] = 'einheit_nr'; ws['B8'] = 'bezeichnung'; ws['C8'] = 'wert'
        ws['A9'] = 'WE01'; ws['B9'] = 'EG links'; ws['C9'] = 65.5
        buf = io.BytesIO(); wb.save(buf)
        datei = SimpleUploadedFile(
            'VS_100007_001.xlsx', buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        resp = self.client.post(
            f'/api/v1/objekte/{self.obj.id}/verteiler/import/preview/',
            {'datei': datei},
            format='multipart',
        )
        self.assertEqual(resp.status_code, 200)
        token = resp.json()['preview_token']
        self.assertIsNotNone(token)

        resp2 = self.client.post(
            f'/api/v1/objekte/{self.obj.id}/verteiler/import/commit/',
            {'preview_token': token},
            format='json',
        )
        self.assertEqual(resp2.status_code, 200)
        self.assertEqual(resp2.json()['anzahl_aktualisiert'], 1)

    def test_protokoll_leer_am_anfang(self):
        resp = self.client.get(f'/api/v1/objekte/{self.obj.id}/verteiler/protokoll/')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json(), [])
