"""Massenimport WEG-Objekte — Parsing, Validierung, Commit."""
from __future__ import annotations

import re
from datetime import date
from io import BytesIO
from typing import Any
from uuid import uuid4

from django.db import transaction
from django.utils import timezone

# ---------------------------------------------------------------------------
# Konstanten
# ---------------------------------------------------------------------------
PLACEHOLDER_IBAN = 'DE00000000000000000000'
AKTUELLES_JAHR   = date.today().year

STANDARD_FREIGABE_GRENZEN = [
    {'bis': 500,   'rolle': 'auto',             'frist_tage': 0, 'beschreibung': 'Automatische Freigabe'},
    {'bis': 5000,  'rolle': 'objektmanager',    'frist_tage': 3, 'beschreibung': 'Objektmanager-Freigabe'},
    {'bis': None,  'rolle': 'geschaeftsfuehrer', 'frist_tage': 5, 'beschreibung': 'Geschäftsführer-Freigabe'},
]

STANDARD_KONTEN_ANZAHL = 70   # Musterkontenrahmen WEG
_PLZ_RE = re.compile(r'^\d{5}$')

_ROMAN_TABLE = [
    (1000, 'M'), (900, 'CM'), (500, 'D'), (400, 'CD'),
    (100,  'C'), (90,  'XC'), (50,  'L'), (40,  'XL'),
    (10,   'X'), (9,   'IX'), (5,   'V'), (4,   'IV'), (1, 'I'),
]


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------
def _roman(n: int) -> str:
    result = ''
    for v, s in _ROMAN_TABLE:
        while n >= v:
            result += s
            n -= v
    return result


def ruecklage_suffix(reihenfolge: int) -> str:
    assert 1 <= reihenfolge <= 21, f'Maximal 21 Rücklagen, got {reihenfolge}'
    suffix = 910 + reihenfolge
    assert suffix != 910, 'Suffix .910 darf niemals vergeben werden'
    return f'.{suffix}'


def _ruecklage_konten_defs(reihenfolge: int) -> list[dict]:
    """5 Sachkonten je Rücklage gemäß Spezifikation Kap. 4.2."""
    s          = 910 + reihenfolge
    roman      = _roman(reihenfolge)
    suffix_str = f'{s % 100:02d}'
    return [
        {
            'kontonummer':      f'099{suffix_str}',
            'kontoname':        f'Rücklagenbestandskonto {roman}',
            'abrechnungsart':   None,
            'direktes_buchen':  False,
            'verteilerschluessel': None,
            'kontoart':         'standard',
            'arge_konto':       False,
        },
        {
            'kontonummer':      f'189{suffix_str}',
            'kontoname':        f'Bank {reihenfolge + 1} Rücklage {roman}',
            'abrechnungsart':   None,
            'direktes_buchen':  False,
            'verteilerschluessel': None,
            'kontoart':         'standard',
            'arge_konto':       False,
        },
        {
            'kontonummer':      f'419{suffix_str}',
            'kontoname':        f'Erlöse Rücklage {roman}',
            'abrechnungsart':   str(s),
            'direktes_buchen':  False,
            'verteilerschluessel': None,
            'kontoart':         'standard',
            'arge_konto':       False,
        },
        {
            'kontonummer':      f'499{suffix_str}',
            'kontoname':        f'Erlöse Entnahme IHR {roman}',
            'abrechnungsart':   '900',
            'direktes_buchen':  False,
            'verteilerschluessel': None,
            'kontoart':         'standard',
            'arge_konto':       False,
        },
        {
            'kontonummer':      f'579{suffix_str}',
            'kontoname':        f'Rücklage {roman} (Aufwand)',
            'abrechnungsart':   str(s),
            'direktes_buchen':  False,
            'verteilerschluessel': '010',
            'kontoart':         'standard',
            'arge_konto':       False,
        },
    ]


def _berechne_konten(anz_rl: int) -> int:
    return STANDARD_KONTEN_ANZAHL + anz_rl * 5


def _berechne_abrechnungsarten(anz_rl: int) -> int:
    # 6 Standard (inkl. 911) + zusätzliche Rücklagen ab 912
    return 6 + max(0, anz_rl - 1)


# ---------------------------------------------------------------------------
# Excel-Vorlage erzeugen
# ---------------------------------------------------------------------------
def erzeuge_vorlage() -> bytes:
    """Erzeugt die MI-WEG.xlsx-Vorlage als Byte-String."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tabelle1'

    headers = ['Objektart', 'Bezeichnung', 'Anschrift 1', 'PLZ1', 'ORT1']
    for i in range(2, 11):
        headers += [f'Anschrift {i}', f'PLZ{i}', f'ORT{i}']
    headers += ['Baujahr', 'ANZ-RL', 'WJ-Jahr', 'WJ-Beginn-Monat']

    pflicht_fill  = PatternFill('solid', fgColor='FFD700')   # gelb
    optional_fill = PatternFill('solid', fgColor='C6EFCE')   # grün
    sonder_fill   = PatternFill('solid', fgColor='BDD7EE')   # blau
    bold = Font(bold=True)

    PFLICHT = {'Objektart', 'Bezeichnung', 'Anschrift 1', 'PLZ1', 'ORT1', 'ANZ-RL'}

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = bold
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
        if header in PFLICHT:
            cell.fill = pflicht_fill
        elif 'Anschrift' in header or header.startswith('PLZ') or header.startswith('ORT'):
            cell.fill = optional_fill
        else:
            cell.fill = sonder_fill

    # Beispielzeile
    total = len(headers)
    ws.cell(row=2, column=1, value='WEG')
    ws.cell(row=2, column=2, value='WEG Musterstraße 1-3')
    ws.cell(row=2, column=3, value='Musterstraße 1')
    ws.cell(row=2, column=4, value='12345')
    ws.cell(row=2, column=5, value='Musterstadt')
    ws.cell(row=2, column=6, value='Musterstraße 3')
    ws.cell(row=2, column=7, value='12345')
    ws.cell(row=2, column=8, value='Musterstadt')
    ws.cell(row=2, column=total - 3, value=1985)     # Baujahr
    ws.cell(row=2, column=total - 2, value=1)        # ANZ-RL
    ws.cell(row=2, column=total - 1, value=AKTUELLES_JAHR)  # WJ-Jahr
    ws.cell(row=2, column=total,     value=1)        # WJ-Beginn-Monat

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Excel parsen
# ---------------------------------------------------------------------------
def _cell_str(val: Any) -> str:
    return '' if val is None else str(val).strip()


def parse_excel(file_bytes: bytes) -> list[dict]:
    """Liest Datenzeilen aus MI-WEG.xlsx. Gibt rohe Zeilendicts zurück."""
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb['Tabelle1'] if 'Tabelle1' in wb.sheetnames else wb.active

    header_row = [_cell_str(ws.cell(row=1, column=c).value)
                  for c in range(1, ws.max_column + 1)]

    required = {'Objektart', 'Bezeichnung', 'Anschrift 1', 'PLZ1', 'ORT1', 'ANZ-RL'}
    missing  = required - set(header_row)
    if missing:
        raise ValueError(f'Fehlende Spalten in der Vorlage: {", ".join(sorted(missing))}')

    def col(name: str) -> int | None:
        try:
            return header_row.index(name)
        except ValueError:
            return None

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        vals = [ws.cell(row=row_idx, column=c).value
                for c in range(1, ws.max_column + 1)]

        if all(v is None or _cell_str(v) == '' for v in vals):
            continue

        def get(name: str) -> str:
            c = col(name)
            return _cell_str(vals[c]) if c is not None and c < len(vals) else ''

        def get_raw(name: str):
            c = col(name)
            return vals[c] if c is not None and c < len(vals) else None

        eingaenge = []
        for i in range(1, 11):
            a_key = 'Anschrift 1' if i == 1 else f'Anschrift {i}'
            p_key = 'PLZ1'        if i == 1 else f'PLZ{i}'
            o_key = 'ORT1'        if i == 1 else f'ORT{i}'
            a, p, o = get(a_key), get(p_key), get(o_key)
            if a or p or o:
                eingaenge.append({'strasse': a, 'plz': p, 'ort': o})
            elif i == 1:
                eingaenge.append({'strasse': '', 'plz': '', 'ort': ''})

        baujahr = None
        bj_raw  = get_raw('Baujahr')
        if bj_raw is not None and _cell_str(bj_raw):
            try:
                baujahr = int(float(str(bj_raw)))
            except (ValueError, TypeError):
                pass

        anz_rl  = None
        rl_raw  = get_raw('ANZ-RL')
        if rl_raw is not None and _cell_str(rl_raw):
            try:
                anz_rl = int(float(str(rl_raw)))
            except (ValueError, TypeError):
                pass

        parse_warnungen = []

        wj_jahr = None
        wj_raw  = get_raw('WJ-Jahr')
        if wj_raw is not None and _cell_str(wj_raw):
            try:
                wj_jahr = int(float(str(wj_raw)))
            except (ValueError, TypeError):
                pass
        if wj_jahr is None:
            wj_jahr = AKTUELLES_JAHR
            parse_warnungen.append(
                f'WJ-Jahr nicht angegeben — auf {AKTUELLES_JAHR} gesetzt.'
            )

        wj_beginn_monat = None
        wbm_raw = get_raw('WJ-Beginn-Monat')
        if wbm_raw is not None and _cell_str(wbm_raw):
            try:
                wj_beginn_monat = int(float(str(wbm_raw)))
            except (ValueError, TypeError):
                pass
        if wj_beginn_monat is None:
            wj_beginn_monat = 1

        rows.append({
            'zeilennummer':    row_idx,
            'objektart':       get('Objektart').upper(),
            'bezeichnung':     get('Bezeichnung'),
            'eingaenge':       eingaenge,
            'baujahr':         baujahr,
            'anz_rl':          anz_rl,
            'wj_jahr':         wj_jahr,
            'wj_beginn_monat': wj_beginn_monat,
            'parse_warnungen': parse_warnungen,
        })

    return rows


# ---------------------------------------------------------------------------
# Validierung
# ---------------------------------------------------------------------------
def validate_zeile(zeile: dict) -> tuple[str, list[str]]:
    """Gibt (status, meldungen) zurück. status in ('ok', 'warnung', 'fehler')."""
    fehler    = []
    warnungen = []

    objektart = zeile.get('objektart', '').strip()
    if not objektart:
        fehler.append('Objektart ist Pflicht (WEG, ZH oder SEV).')
    elif objektart not in ('WEG', 'ZH', 'SEV'):
        fehler.append(f'Objektart "{objektart}" ungültig — erlaubt: WEG, ZH, SEV.')

    bez = zeile.get('bezeichnung', '').strip()
    if not bez:
        fehler.append('Bezeichnung ist Pflicht.')
    elif len(bez) > 200:
        fehler.append('Bezeichnung darf max. 200 Zeichen haben.')

    eingaenge = zeile.get('eingaenge', [])
    if eingaenge:
        e1 = eingaenge[0]
        if not e1.get('strasse'):
            fehler.append('Anschrift 1 ist Pflicht.')
        if not e1.get('plz'):
            fehler.append('PLZ1 ist Pflicht.')
        elif not _PLZ_RE.match(e1['plz']):
            fehler.append('PLZ1 muss eine 5-stellige Zahl sein.')
        if not e1.get('ort'):
            fehler.append('ORT1 ist Pflicht.')

    for i, e in enumerate(eingaenge[1:], start=2):
        a, p, o = e.get('strasse', ''), e.get('plz', ''), e.get('ort', '')
        if a or p or o:
            if not a:
                fehler.append(f'Eingang {i}: Anschrift fehlt (PLZ/ORT sind gefüllt).')
            if not p:
                fehler.append(f'Eingang {i}: PLZ fehlt.')
            elif not _PLZ_RE.match(p):
                fehler.append(f'Eingang {i}: PLZ muss 5-stellig sein.')
            if not o:
                fehler.append(f'Eingang {i}: Ort fehlt.')

    baujahr = zeile.get('baujahr')
    if baujahr is None:
        warnungen.append('Baujahr fehlt.')
    elif not (1800 <= baujahr <= AKTUELLES_JAHR):
        fehler.append(f'Baujahr muss zwischen 1800 und {AKTUELLES_JAHR} liegen.')

    anz_rl = zeile.get('anz_rl')
    if anz_rl is None:
        fehler.append('ANZ-RL ist Pflicht.')
    elif not isinstance(anz_rl, int) or not (0 <= anz_rl <= 21):
        fehler.append('ANZ-RL muss eine ganze Zahl zwischen 0 und 21 sein.')

    for w in zeile.get('parse_warnungen', []):
        warnungen.append(w)

    wj_jahr = zeile.get('wj_jahr')
    if wj_jahr is not None and not (2000 <= wj_jahr <= AKTUELLES_JAHR + 1):
        fehler.append(f'WJ-Jahr muss zwischen 2000 und {AKTUELLES_JAHR + 1} liegen.')

    wj_beginn_monat = zeile.get('wj_beginn_monat')
    if wj_beginn_monat is not None and not (1 <= wj_beginn_monat <= 12):
        fehler.append('WJ-Beginn-Monat muss zwischen 1 und 12 liegen.')

    adressen = [f"{e['strasse']}|{e['plz']}|{e['ort']}"
                for e in eingaenge if e.get('strasse')]
    if len(adressen) != len(set(adressen)):
        warnungen.append('Mehrere Eingänge mit identischer Anschrift.')

    if fehler:
        return 'fehler', fehler + warnungen
    if warnungen:
        return 'warnung', warnungen
    return 'ok', []


# ---------------------------------------------------------------------------
# Vorschau erstellen (kein DB-Commit der Objekte)
# ---------------------------------------------------------------------------
def preview_erstellen(file_bytes: bytes, user) -> dict:
    from .models import ImportJob
    from apps.objekte.models import Objekt

    rows = parse_excel(file_bytes)
    if len(rows) > 500:
        raise ValueError('Maximal 500 Zeilen pro Import-Datei.')

    bestehende_bezeichnungen = set(
        Objekt.objects.values_list('bezeichnung', flat=True)
    )

    zeilen_preview = []
    for zeile in rows:
        status, meldungen = validate_zeile(zeile)

        # Warnung bei doppelter Bezeichnung
        bez = zeile.get('bezeichnung', '')
        if bez and bez in bestehende_bezeichnungen and status != 'fehler':
            status = 'warnung'
            meldungen.insert(0, f'Bezeichnung "{bez}" existiert bereits (Duplikat-Warnung).')

        eingaenge_anzahl = sum(
            1 for e in zeile.get('eingaenge', []) if e.get('strasse')
        )
        anz_rl = zeile.get('anz_rl') or 0

        zeilen_preview.append({
            'zeilennummer':            zeile['zeilennummer'],
            'status':                  status,
            'meldungen':               meldungen,
            'bezeichnung':             bez,
            'eingaenge_anzahl':        eingaenge_anzahl,
            'ruecklagen':              anz_rl,
            'konten_anzahl':           _berechne_konten(anz_rl),
            'abrechnungsarten_anzahl': _berechne_abrechnungsarten(anz_rl),
            'wj_jahr':                 zeile.get('wj_jahr'),
            'wj_beginn_monat':         zeile.get('wj_beginn_monat'),
            '_daten':                  zeile,
        })

    ok      = sum(1 for z in zeilen_preview if z['status'] == 'ok')
    warnung = sum(1 for z in zeilen_preview if z['status'] == 'warnung')
    fehler  = sum(1 for z in zeilen_preview if z['status'] == 'fehler')

    importierbar = [z for z in zeilen_preview if z['status'] in ('ok', 'warnung')]
    summary = {
        'ok':               ok,
        'warnung':          warnung,
        'fehler':           fehler,
        'gesamt':           len(zeilen_preview),
        'objekte':          len(importierbar),
        'liegenschaften':   sum(z['eingaenge_anzahl'] for z in importierbar),
        'bankkonten':       sum(1 + z['ruecklagen'] for z in importierbar),
        'konten':           sum(z['konten_anzahl'] for z in importierbar),
        'abrechnungsarten': sum(z['abrechnungsarten_anzahl'] for z in importierbar),
    }

    token = uuid4()
    job   = ImportJob.objects.create(
        typ='weg_objekt',
        datei_pfad='',
        status='parsed',
        preview_token=token,
        zeilen_gesamt=len(zeilen_preview),
        zeilen_ok=ok,
        zeilen_warnung=warnung,
        zeilen_fehler=fehler,
        ergebnis={'zeilen': zeilen_preview, 'summary': summary},
        erstellt_von=user,
    )

    return {
        'job_id':        str(job.id),
        'preview_token': str(token),
        'zeilen':        zeilen_preview,
        'summary':       summary,
    }


# ---------------------------------------------------------------------------
# Einzelnes Objekt importieren (atomar)
# ---------------------------------------------------------------------------
@transaction.atomic
def _importiere_eine_zeile(zeile: dict, user) -> dict:
    from apps.objekte.models import Objekt, Eingang, Bankkonto, Wirtschaftsjahr
    from apps.konten.models import Konto, Abrechnungsart
    from apps.konten.services import kontenrahmen_anlegen, abrechnungsarten_anlegen, verteilerschluessel_anlegen

    anz_rl          = zeile.get('anz_rl') or 0
    eingaenge       = zeile.get('eingaenge', [])
    e1              = eingaenge[0] if eingaenge else {}
    objektart       = zeile.get('objektart', 'WEG')
    wj_jahr         = zeile.get('wj_jahr') or AKTUELLES_JAHR
    wj_beginn_monat = zeile.get('wj_beginn_monat') or 1

    # 1. Objekt
    objekt = Objekt.objects.create(
        objekt_typ=objektart,
        bezeichnung=zeile['bezeichnung'],
        strasse=e1.get('strasse', ''),
        plz=e1.get('plz', ''),
        ort=e1.get('ort', ''),
        baujahr=zeile.get('baujahr'),
        verwaltung_seit=date.today(),
        wirtschaftsjahr_start=wj_beginn_monat,
        zahlungsfreigabe_grenzen=STANDARD_FREIGABE_GRENZEN,
        status='aktiv',
    )

    # 2a. Erstes Wirtschaftsjahr
    wj = Wirtschaftsjahr.objects.create(
        objekt=objekt,
        jahr=wj_jahr,
        beginn_monat=wj_beginn_monat,
        status='offen',
        eroeffnet_von=user,
    )

    # 2b. Eingänge / Liegenschaften
    for i, e in enumerate(eingaenge, start=1):
        if not e.get('strasse'):
            continue
        Eingang.objects.create(
            objekt=objekt,
            bezeichnung=e['strasse'],
            strasse=e['strasse'],
            plz=e['plz'],
            ort=e['ort'],
        )

    # 3. Bewirtschaftungskonto (Platzhalter)
    Bankkonto.objects.create(
        objekt=objekt,
        konto_typ='bewirtschaftung',
        bezeichnung='Bewirtschaftung',
        iban=PLACEHOLDER_IBAN,
        bic='',
        kontoinhaber=objekt.bezeichnung,
        reihenfolge=0,
        aktiv=False,
    )

    # 4. Rücklagen-Bankkonten (Platzhalter)
    for n in range(1, anz_rl + 1):
        Bankkonto.objects.create(
            objekt=objekt,
            konto_typ='ruecklage',
            bezeichnung=f'Rücklage {_roman(n)}',
            iban=PLACEHOLDER_IBAN,
            bic='',
            kontoinhaber=objekt.bezeichnung,
            reihenfolge=n,
            aktiv=False,
        )

    # 5. Standard-Kontenrahmen (enthält Rücklage-I-Konten bereits) — nur für WEG
    if objektart == 'WEG':
        kontenrahmen_anlegen(wirtschaftsjahr_id=str(wj.id))

    # 6. Dynamische Rücklagen-Konten (für alle N, get_or_create ist idempotent) — nur WEG
    if objektart == 'WEG':
        for n in range(1, anz_rl + 1):
            for kd in _ruecklage_konten_defs(n):
                Konto.objects.get_or_create(
                    wirtschaftsjahr=wj,
                    kontonummer=kd['kontonummer'],
                    defaults={**kd, 'arge_kostenart': None, 'aktiv': True},
                )

    # 7. Standard-Abrechnungsarten (inkl. 911 für Rücklage I)
    abrechnungsarten_anlegen(str(objekt.id))

    # 8. Zusatz-Abrechnungsarten für Rücklage II+
    for n in range(2, anz_rl + 1):
        Abrechnungsart.objects.get_or_create(
            objekt=objekt,
            code=str(910 + n),
            defaults={'bezeichnung': f'Rücklage {_roman(n)}', 'aktiv': True},
        )

    # 9. Standard-Verteilerschlüssel (Einheiten noch nicht vorhanden — Werte folgen beim Einheiten-Import)
    verteilerschluessel_anlegen(str(objekt.id))

    return {'objekt_id': str(objekt.id), 'bezeichnung': objekt.bezeichnung}


# ---------------------------------------------------------------------------
# Commit
# ---------------------------------------------------------------------------
def commit_import(preview_token: str, user) -> dict:
    from .models import ImportJob

    try:
        job = ImportJob.objects.get(preview_token=preview_token, status='parsed')
    except ImportJob.DoesNotExist:
        raise ValueError('Vorschau-Token ungültig oder bereits verwendet.')

    # 30-Minuten-Ablauf
    alter_sek = (timezone.now() - job.erstellt_am).total_seconds()
    if alter_sek > 1800:
        raise ValueError('Vorschau-Token abgelaufen (30-Minuten-Frist überschritten).')

    zeilen     = job.ergebnis.get('zeilen', [])
    ergebnisse = []
    ok_count   = 0
    err_count  = 0

    for z in zeilen:
        if z['status'] == 'fehler':
            ergebnisse.append({
                'zeilennummer': z['zeilennummer'],
                'status':       'uebersprungen',
                'meldung':      'Validierungsfehler — Zeile nicht importiert.',
            })
            continue
        try:
            result = _importiere_eine_zeile(z['_daten'], user)
            ergebnisse.append({
                'zeilennummer': z['zeilennummer'],
                'status':       'ok',
                'objekt_id':    result['objekt_id'],
                'bezeichnung':  result['bezeichnung'],
            })
            ok_count += 1
        except Exception as exc:
            ergebnisse.append({
                'zeilennummer': z['zeilennummer'],
                'status':       'fehler',
                'meldung':      str(exc),
            })
            err_count += 1

    if ok_count == 0:
        new_status = 'failed'
    elif err_count > 0:
        new_status = 'partial'
    else:
        new_status = 'committed'

    job.status  = new_status
    job.ergebnis = {**job.ergebnis, 'commit_ergebnisse': ergebnisse}
    job.save(update_fields=['status', 'ergebnis', 'aktualisiert_am'])

    return {
        'job_id':    str(job.id),
        'status':    new_status,
        'importiert': ok_count,
        'fehler':    err_count,
        'ergebnisse': ergebnisse,
    }
